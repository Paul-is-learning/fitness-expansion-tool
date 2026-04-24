#!/usr/bin/env python3
"""
excel_to_ir.py — Parse le fichier Excel BP FP Romania en IR JSON.

IR format (stable, source de vérité pour engine.js):
{
  "source": "<filename>",
  "generated_at": "<iso>",
  "sheets_order": ["EXEC_SUMMARY", "HYPOTHESES", ...],
  "sheets": {
    "HYPOTHESES": {
      "dimensions": "A1:Q206",
      "max_row": 206,
      "max_col": 17,
      "cells": {
        "A1": {"t": "s", "v": "BP MASTER..."},
        "C14": {"t": "n", "v": 45000},
        "C15": {"t": "f", "f": "=C16*6", "v_excel": 0.06},
        ...
      }
    }
  },
  "named_ranges": [],
  "input_candidates": {
    "HYPOTHESES": ["C14", "C15", ...]   // constantes littérales col C lignes 14-100
  }
}

Règle critique : on stocke la FORMULE (f) + la VALEUR CALCULÉE PAR EXCEL (v_excel).
engine.js doit retomber sur v_excel (tolérance 0.01) pour chaque cell → c'est
le golden test.
"""
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

FILE = Path("/Users/paulbecaud/Desktop/VF MF FP/MF FP - BP RO - v2Financement mix.xlsx")
OUT = Path("/Users/paulbecaud/Desktop/fitness-expansion-tool/src/bp/bp_ir.json")


def cell_to_entry(cell, cell_val):
    """Convertit une cellule openpyxl en entrée IR.

    t = type : 's' string | 'n' number | 'b' bool | 'd' date | 'f' formula | 'e' error | 'z' empty
    v = valeur brute (string/number/bool/date ISO)
    f = formule string (si t='f')
    v_excel = valeur calculée par Excel (pour formules, source de vérité golden)
    """
    v = cell.value
    if v is None:
        return None  # On skippe les vides pour alléger l'IR

    # Formule
    if isinstance(v, str) and v.startswith("="):
        val_excel = cell_val.value
        # Heuristique : cellule-note (texte en format Excel Text) qui commence par "= "
        # suivie de caractères manifestement non-formule (× ÷ €). Dans ce cas Excel
        # renvoie la chaîne telle quelle (val_excel == v) ou None. On déclasse en 's'.
        # Détection robuste : si Excel n'a pas évalué la formule (val_excel == v,
        # c-à-d Excel renvoie la chaîne brute), c'est une cellule formattée en texte.
        # Ne pas la traiter comme formule pour éviter des faux positifs.
        if val_excel is not None and isinstance(val_excel, str) and val_excel == v:
            return {"t": "s", "v": v}
        body = v[1:].lstrip()
        has_non_formula_chars = any(ch in body for ch in ('×', '÷', '€', '—', '–'))
        if has_non_formula_chars and val_excel is None:
            return {"t": "s", "v": v}
        return {
            "t": "f",
            "f": v,
            "v_excel": val_excel,
        }

    # String
    if isinstance(v, str):
        return {"t": "s", "v": v}

    # Bool (avant int — Python bool est int)
    if isinstance(v, bool):
        return {"t": "b", "v": v}

    # Nombre
    if isinstance(v, (int, float)):
        return {"t": "n", "v": v}

    # Date
    try:
        from datetime import datetime as _dt, date as _d
        if isinstance(v, (_dt, _d)):
            return {"t": "d", "v": v.isoformat()}
    except Exception:
        pass

    # Fallback : on stocke en string
    return {"t": "s", "v": str(v)}


def main():
    if not FILE.exists():
        sys.exit(f"Fichier introuvable: {FILE}")

    print(f"[IR] Parsing {FILE.name}...", file=sys.stderr)
    wb = load_workbook(FILE, data_only=False)
    wb_val = load_workbook(FILE, data_only=True)

    ir = {
        "source": FILE.name,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "sheets_order": list(wb.sheetnames),
        "sheets": {},
        "named_ranges": [],
        "input_candidates": {"HYPOTHESES": []},
    }

    # Named ranges (attendu 0 d'après audit)
    try:
        for name in wb.defined_names:
            dn = wb.defined_names[name]
            ir["named_ranges"].append({"name": name, "value": str(dn.value)})
    except Exception as e:
        ir["named_ranges_error"] = str(e)

    total_formulas = 0
    total_cells = 0

    for ws in wb.worksheets:
        ws_val = wb_val[ws.title]
        sheet_cells = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cell_val = ws_val[cell.coordinate]
                entry = cell_to_entry(cell, cell_val)
                if entry is None:
                    continue
                sheet_cells[cell.coordinate] = entry
                total_cells += 1
                if entry["t"] == "f":
                    total_formulas += 1

        ir["sheets"][ws.title] = {
            "state": ws.sheet_state,
            "dimensions": ws.dimensions,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "cells": sheet_cells,
        }

    # Input candidates — constantes littérales HYPOTHESES col C, lignes 14..100
    if "HYPOTHESES" in ir["sheets"]:
        hyp = ir["sheets"]["HYPOTHESES"]["cells"]
        for r in range(14, 101):
            coord = f"C{r}"
            entry = hyp.get(coord)
            if entry and entry["t"] in ("n", "s", "b"):
                # on exclut les formules (calculs dérivés) — seules les constantes sont inputs
                ir["input_candidates"]["HYPOTHESES"].append(coord)

    ir["stats"] = {
        "total_cells": total_cells,
        "total_formulas": total_formulas,
        "input_candidates_count": len(ir["input_candidates"]["HYPOTHESES"]),
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w") as fh:
        json.dump(ir, fh, indent=2, ensure_ascii=False, default=str)

    print(f"[IR] {total_cells} cells, {total_formulas} formulas, {len(ir['input_candidates']['HYPOTHESES'])} input candidates", file=sys.stderr)
    print(f"[IR] Wrote {OUT} ({OUT.stat().st_size} bytes)", file=sys.stderr)


if __name__ == "__main__":
    main()
